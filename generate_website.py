"""
generate_website.py
Reads stallion_data.xlsx → generates index.html
Single-page responsive website with Darley brand colours.

Usage:  python generate_website.py
"""

import re
import base64
import io
import urllib.request
import urllib.parse
from html.parser import HTMLParser
from pathlib import Path
from collections import defaultdict
from datetime import date 
today = date.today()

BASE   = Path(r"C:\Users\bsharp\OneDrive - Godolphin\PythonScripts\ClaudeCodeProjects\cheat-sheets")
XL     = BASE / "stallion_data.xlsx"
OUT    = BASE / "index.html"
PP_DIR = BASE / "stallion PPs png files"

# ── Full-size conformation photo URLs from darleyamerica.com ──────────────────
CDN_BASE = "https://cdn.darleystallions.com/sites/default/files/drupal-media/stallion-images/"
PHOTO_URLS = {
    "nyquist.jpg":           "https://cdn.darleystallions.com/sites/default/files/drupal-media/stallion-images/USA-2016/nyquist/conf_nyquist_thoroughbred_stallion.jpg",
    "codys_wish.jpg":        CDN_BASE + "USA-2019/Cody's%20Wish/conf_Codys_Wish_250717_CN_MW-2965-RT3_e_thoroughbred_stallion.jpg",
    "maxfield.jpg":          CDN_BASE + "USA-2019/Maxfield/Conf_Maxfield_220818_CN_AL-0012-RT_98_thoroughbred_stallions.jpg",
    "street_sense.jpg":      "https://cdn.darleystallions.com/sites/default/files/drupal-media/stallion-images/USA/street-sense/st-sense-739-5021.jpg",
    "essential_quality.jpg": CDN_BASE + "USA-2019/Essential%20Quality/CONF_Essential_Quality_240917_CN_MR_0189-RT_98_thoroughbred_stallion.jpg",
    "hard_spun.jpg":         CDN_BASE + "USA-2019/hard-spun/conf_Hard_Spun_0095-RT_thoroughbred_stallion.jpg",
    "midshipman.jpg":        "https://cdn.darleystallions.com/sites/default/files/conformations/Midshipman-Conf.jpg",
    "frosted.jpg":           "https://cdn.darleystallions.com/sites/default/files/drupal-media/stallion-images/USA-2016/frosted/conf_frosted_thoroughbred_stallion.jpg",
    "proxy.jpg":             CDN_BASE + "USA-2019/Proxy/conf_Proxy_251002_CN_MW-3926-RT_e_thoroughbred_stallion.jpg",
    "speakers_corner.jpg":   CDN_BASE + "USA-2019/Speaker's%20Corner/CONF_Speakers_Corner_240916_CN_MR_0200-RT_98_thoroughbred_stallion.jpg",
    "first_mission.jpg":     CDN_BASE + "USA-2019/First%20Mission/conf_FIRST_MISSION_CDY_Essex_H_032324_Fini_1_98_throughbred_stallion.jpg",
    "highland_falls.jpg":    CDN_BASE + "USA-2019/Highland%20Falls/conf_Highland_falls_AC-jcgc_sr_98_thoroughbred_stallion.jpg",
    "mystic_guide.jpg":      CDN_BASE + "USA-2019/Mystic%20Guide/CONF_Mystic_Guide_240918_CN_MR_0259-RT_98_thoroughbred_stallion.jpg",
}

# ── Selling points fallback (used if live scrape fails) ──────────────────────
_SP_FALLBACK = {
    "Nyquist": [
        "Three Graded Stakes winners already in 2026 - only Not This Time and Into Mischief have more.",
        "Four G1 winners and 10 G1 horses in 2025 - only Into Mischief had more.",
        "His 4% G1 horses from starters in 2025 is best among all stallions.",
        "12 Graded Stakes winners in 2025 - best among all stallions under $250k.",
        "A $1.2m juvenile at OBS March 2026, and 2025 yearlings bring $1.05m, $1m, $850k, $750k, $735k, $725k, $725k.",
    ],
    "Cody's Wish": [
        "The 2023 Horse of the Year.",
        "More G1 wins at a mile on dirt than any other colt ever.",
        "Sensational two-time winner of the G1 Breeders' Cup Dirt Mile, plus the G1 Met Mile, G1 Churchill Downs Stakes, and G1 Forego.",
        "Curlin's only five-time G1-winning son.",
        "His 112 Beyer in the G1 Forego makes him Curlin's fastest horse at less than a mile.",
        "Eight career triple-digit Beyers.",
        "His 112 in the G1 Met Mile was the highest of 2023.",
        "Out of a G1 winner by Tapit and from the La Troienne family, like roster mates Essential Quality and Maxfield.",
        "First weanlings in 2025 bring $550k, $500k, $440k, $425k, $425k.",
    ],
    "Maxfield": [
        "12 first-crop Stakes horses and four TDN Rising Stars.",
        "His 18% Stakes horses from starters is best among all stallions under $100k.",
        "First-crop two-year-olds bring $1.25m, $1m, \u20ac1m. Plus two $1m first-crop yearlings.",
        "Only American Pharoah and Flightline have ever had more first-crop seven-figure sales horses.",
        "Second-crop yearlings in 2025 bring $550k, $475k, $450k, $425k.",
        "First three-year-olds in 2026.",
    ],
    "Street Sense": [
        "Thirteen career G1 winners include four stallion sons in Kentucky.",
        "Maxfield - freshman sensation.",
        "McKinzie - already the sire of three G1 winners from his first crop.",
        "Speaker's Corner - whose 114 Beyer in the G1 Carter is the highest in the past six years.",
        "First Mission - multiple Graded Stakes winner.",
        "Bella Ballerina - G2 Rachel Alexandra winner and Kentucky Oaks contender.",
        "La Cara - G1 Ashland and G1 Acorn winner.",
        "Street Beast - $1m Kentucky Downs Juvenile Mile winner.",
        "2025 yearlings bring $650k, $600k, $500k, $450k, $425k, $400k.",
    ],
    "Essential Quality": [
        "Five first-crop Stakes winners.",
        "Leading second-crop sire by earnings.",
        "Only Gun Runner and Into Mischief can beat his three three-year-old Stakes winners in 2026.",
        "Tapit's leading money-earner.",
        "The only Breeders' Cup Juvenile winner ever to capture the Belmont Stakes.",
        "Only the seventh colt in the 53-year history of the Eclipse Awards to be named Champion at both two and three.",
        "First three-year-olds in 2026.",
    ],
    "Hard Spun": [
        "Fifteen career G1 winners and four stallion sons in Kentucky: G1 Met Mile winner Silver State, G1 Breeders' Cup Sprint hero Aloha West, G1 Breeders' Cup Dirt Mile winner Spun to Run, and Kentucky Derby runner-up Two Phil's.",
        "The leading sire of career Stakes horses among active stallions under $60,000.",
        "21 Black Type horses in 2025 - more than any other stallion under $45k - including Graded Stakes winners Pondering, Gas Me Up and Elysian Field.",
        "2025 juveniles who brought $335k, $305k and $230k show he's very much in demand.",
    ],
    "Midshipman": [
        "His 10% juvenile Stakes winners from runners in 2025 can only be topped by Gun Runner, Not This Time and Upstart.",
        "His 17% career Black Type horses from starters continues to be among the best in the business - no stallion under $100k can top it.",
        "A $650k juvenile at OBS March and 2025 yearlings include a $270k session-topper at Keeneland.",
    ],
    "Frosted": [
        "The fastest-ever winner of the stallion-making G1 Met Mile, running an astonishing 123 Beyer.",
        "The best figure ever recorded at a mile.",
        "Eight Black Type winners in 2025.",
        "Post Time's 110 Beyer was among the best of the year - only five horses ran faster in 2025.",
        "Royal Ascot G1 King Charles III Stakes runner-up Frost at Dawn.",
        "G3 Gotham winner Flood Zone.",
        "$1 million Kentucky Downs Juvenile Fillies Stakes winner Meringue.",
    ],
    "Proxy": [
        "Outstanding good looks and Tapit's only G1-winning son out of a multiple G1-winning mare.",
        "Winner of the G1 Clark at Churchill Downs and the $1 million G2 Oaklawn Handicap, plus a fast-closing third in the G1 Breeders' Cup Classic.",
        "Won or placed in 11 Graded Stakes, including the G1 Breeders' Cup Classic, G1 Jockey Club Gold Cup and G1 Santa Anita Handicap, and defeated 22 G1 winners and an additional 23 Graded Stakes winners.",
    ],
    "Speaker's Corner": [
        "The fastest ever by sire of sires Street Sense. Earned a 114 Beyer in the G1 Carter.",
        "In the past six years only Sovereignty and Flightline have run faster.",
        "His seven career triple-digit Beyers include a 109 at three.",
        "Broke his maiden in an especially hot Belmont MSW at two, beating five future Stakes winners.",
        "Bred on the same cross as star freshman Maxfield - out of a Bernardini daughter of G1 Breeders' Cup Distaff winner Round Pond.",
        "First yearlings in 2025 averaged $66k and brought $250k, $210k, $185k, $170k, $165k, $150k, and four at $140k.",
    ],
    "First Mission": [
        "Won the $1.25m G2 Oaklawn Handicap running a 109 Beyer and beating G1 Pegasus World Cup winner Skippylongstocking.",
        "Plus the G2 Alysheba at Churchill running a 106 Beyer, the G3 Essex Stakes and G3 Lexington Stakes.",
        "His 109 Beyer makes him the fastest freshman to retire for $10k or less in the past five years.",
        "Four Graded Stakes wins and six triple digit Beyers.",
        "Third in the G1 Stephen Foster to Mindframe and Sierra Leone, beating Kentucky Derby winner Mystik Dan and Dubai World Cup winner Hit Show.",
        "By sire of sires Street Sense, whose sons at stud include star freshman Maxfield and McKinzie, already the sire of three G1 winners from his first crop.",
        "Free of A. P. Indy - his sire's best cross.",
    ],
    "Highland Falls": [
        "Brilliant four-length winner of the G1 Jockey Club Gold Cup, like his sire Curlin, running a 104 Beyer and beating three G1 winners.",
        "Runner-up by a length to Sierra Leone in the G1 Whitney, running a 107 Beyer and beating G1 winners Fierceness, White Abarrio and Skippylongstocking.",
        "Curlin's only G1-winning son out of a multiple G1-winning mare, his dam is G1 Breeders' Cup Distaff and G1 Acorn winner Round Pond.",
    ],
    "Mystic Guide": [
        "Rated the best in the world after winning the G1 Dubai World Cup in a faster time than Arrogate and California Chrome.",
        "Also won the G2 Jim Dandy at three and took the G3 Razorback by six lengths earning a 108 Beyer.",
        "Dam is A.P. Indy's five-time G1-winning daughter Music Note, a half-sister to multiple G1 and Classic winner Musical Chimes from the family of Champion two-year-old filly It's In The Air and 2,000 Guineas winner Coroebus.",
        "First yearlings in 2025 averaged $63k and brought $425k, $300k, $245k, $210k, $180k, $150k, $150k, $150k, $140k, $140k.",
    ],
}

# ── Live scraper ─────────────────────────────────────────────────────────────

def _stallion_slug(name):
    return re.sub(r"[^a-z0-9]+", "-", name.lower().replace("'", "")).strip("-")

def _fetch_page(name):
    """Fetch raw HTML for a stallion page. Returns html string or None."""
    slug = _stallion_slug(name)
    url  = f"https://www.darleyamerica.com/stallions/our-stallions/{slug}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            return resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  WARNING: could not fetch {name} ({e})")
        return None

def _parse_selling_points(name, html):
    """Extract selling points from page HTML."""
    class _Parser(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.in_summary = False
            self.summary_depth = -1
            self.div_depth = 0
            self.in_p = False
            self.paragraphs = []
            self._buf = ""

        def handle_starttag(self, tag, attrs):
            cls = dict(attrs).get("class", "")
            if tag == "div":
                self.div_depth += 1
                if "summary" in cls and not self.in_summary:
                    self.in_summary = True
                    self.summary_depth = self.div_depth
            if self.in_summary and tag == "p":
                self.in_p = True
                self._buf = ""
            if self.in_p and tag == "strong":
                self._buf += "<strong>"

        def handle_endtag(self, tag):
            if self.in_p and tag == "strong":
                self._buf += "</strong>"
            if self.in_summary and tag == "p" and self.in_p:
                text = self._buf.strip()
                if text:
                    self.paragraphs.append(text)
                self.in_p = False
            if tag == "div":
                if self.in_summary and self.div_depth == self.summary_depth:
                    self.in_summary = False
                self.div_depth -= 1

        def handle_data(self, data):
            if self.in_p:
                self._buf += data

    _EXCLUDE = ('breeding form', 'stud office', 'nominations team',
                'request a nomination', 'talk to our', 'click here')
    p = _Parser()
    p.feed(html)
    points = [pt for pt in p.paragraphs
              if not any(x in pt.lower() for x in _EXCLUDE)]
    if points:
        return points
    return _SP_FALLBACK.get(name, [])

def _parse_profile(name, html):
    """Extract stallion profile fields from page HTML. Returns a dict."""
    ped = PEDIGREES.get(name, {})
    photo_f = name.lower().replace(' ', '_').replace("'", "") + '.jpg'

    def _find(pattern, text, group=1):
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        return m.group(group).strip() if m else ''

    # Stud fee: look for price pattern near "fee" (first match wins)
    fee = _find(r'(?:stud\s+fee|fee)[^$\n]{0,40}\$([\d,]+)', html)
    if not fee:
        fee = _find(r'\$([\d,]+)', html)
    fee = ('$' + fee) if fee else ''

    # Year foaled
    foaled = _find(r'(?:foaled|year\s+foaled)[^\d]{0,20}(\d{4})', html)

    # Earnings
    earn = _find(r'(?:earnings?|career\s+earnings?)[^$\n]{0,20}\$([\d,.]+(?:\s*million)?)', html)
    earnings = ('$' + earn) if earn else ''

    # Entered stud
    entered = _find(r'(?:entered\s+stud|stud\s+since|standing\s+since)[^\d]{0,20}(\d{4})', html)

    # First crop note
    first_crop = _find(r'(first\s+(?:crop|foals?)[^<.\n]{0,60})', html)

    return {
        'name':           name,
        'fee_current':    fee,
        'sire':           ped.get('sire', ''),
        'dam':            ped.get('dam', ''),
        'damsire':        ped.get('dam_sire', ''),
        'year_foaled':    foaled,
        'height':         '',
        'earnings':       earnings,
        'entered_stud':   entered,
        'first_crop_note': first_crop,
        'photo_file':     photo_f,
    }

def scrape_all_stallions(names):
    """Scrape each stallion page once; return (profiles list, sp_map dict).
    profiles is a list of dicts in the same order as names.
    sp_map is name → [selling points].
    """
    print("Scraping stallion data from darleyamerica.com...")
    profiles = []
    sp_map   = {}
    for name in names:
        print(f"  {name}...", end=" ", flush=True)
        html = _fetch_page(name)
        if html:
            profile = _parse_profile(name, html)
            points  = _parse_selling_points(name, html)
        else:
            profile = _parse_profile(name, "")   # returns blanks + PEDIGREES data
            points  = _SP_FALLBACK.get(name, [])
        profiles.append(profile)
        sp_map[name] = points
        print(f"{len(points)} selling points")
    return profiles, sp_map

# ── Pedigree data (3 generations) ────────────────────────────────────────────
PEDIGREES = {
    "Nyquist":           {"sire":"Uncle Mo","dam":"Seeking Gabrielle","sire_sire":"Indian Charlie","sire_dam":"Political Parfait","dam_sire":"Forestry","dam_dam":"Seeking Regina"},
    "Cody's Wish":       {"sire":"Curlin","dam":"Dance Card","sire_sire":"Smart Strike","sire_dam":"Classy 'N Smart","dam_sire":"Tapit","dam_dam":"Tempting Note"},
    "Maxfield":          {"sire":"Street Sense","dam":"Velvety","sire_sire":"Street Cry","sire_dam":"Bedazzle","dam_sire":"Bernardini","dam_dam":"Cara Rafaela"},
    "Street Sense":      {"sire":"Street Cry","dam":"Bedazzle","sire_sire":"Machiavellian","sire_dam":"Coup de Folie","dam_sire":"Dixieland Band","dam_dam":"Majestic Legend"},
    "Essential Quality": {"sire":"Tapit","dam":"Delightful Quality","sire_sire":"Pulpit","sire_dam":"Tap Your Heels","dam_sire":"Elusive Quality","dam_dam":"Contrive"},
    "Hard Spun":         {"sire":"Danzig","dam":"Turkish Tryst","sire_sire":"Northern Dancer","sire_dam":"Pas de Nom","dam_sire":"Turkoman","dam_dam":"Taba"},
    "Midshipman":        {"sire":"Unbridled's Song","dam":"Fleet Lady","sire_sire":"Unbridled","sire_dam":"Trolley Song","dam_sire":"Avenue of Flags","dam_dam":"Seaside Attraction"},
    "Frosted":           {"sire":"Tapit","dam":"Fast Cookie","sire_sire":"Pulpit","sire_dam":"Tap Your Heels","dam_sire":"Deputy Minister","dam_dam":"Mint Copy"},
    "Proxy":             {"sire":"Tapit","dam":"Panty Raid","sire_sire":"Pulpit","sire_dam":"Tap Your Heels","dam_sire":"Include","dam_dam":"Adventurous Di"},
    "Speaker's Corner":  {"sire":"Street Sense","dam":"Tyburn Brook","sire_sire":"Street Cry","sire_dam":"Bedazzle","dam_sire":"Bernardini","dam_dam":"Round Pond"},
    "First Mission":     {"sire":"Street Sense","dam":"Elude","sire_sire":"Street Cry","sire_dam":"Bedazzle","dam_sire":"Medaglia d'Oro","dam_dam":"Cappucino Bay"},
    "Highland Falls":    {"sire":"Curlin","dam":"Round Pond","sire_sire":"Smart Strike","sire_dam":"Classy 'N Smart","dam_sire":"Awesome Again","dam_dam":"Primal Force"},
    "Mystic Guide":      {"sire":"Ghostzapper","dam":"Music Note","sire_sire":"Awesome Again","sire_dam":"Baby Zip","dam_sire":"A.P. Indy","dam_dam":"Note Musicale"},
}

# ── helpers ───────────────────────────────────────────────────────────────────

def slugify(name):
    return re.sub(r'[^a-z0-9]+', '-', str(name).lower()).strip('-')

def esc(text):
    return (str(text)
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;'))

def fmt_currency(val):
    """Format a value as $X,XXX — works on numbers or plain numeric strings."""
    if val is None or str(val).strip() == '':
        return ''
    s = str(val).strip().lstrip('$').replace(',', '')
    try:
        n = float(s)
        return f'${n:,.0f}'
    except ValueError:
        return str(val)  # already a string like "Private" — return as-is

def _fmt_pct(val):
    """Format a value as X.X% (multiplying by 100) — returns empty string if blank/None."""
    if val is None or str(val).strip() == '' or str(val).strip().lower() == 'none':
        return ''
    try:
        return f'{float(val) * 100:.1f}%'
    except (ValueError, TypeError):
        return str(val)


def load_tables_from_db():
    """Load fee history and sale results from SQL. Returns (fh_map, sr_map, hl_map).
    Highlights not yet in DB — hl_map always loaded from Excel."""
    import config
    engine = config.get_engine()
    def query(sql):
        with engine.connect() as conn:
            result = conn.execute(sql)
            return [dict(row) for row in result.mappings()]

    fee_hist = query(config.SQL_FEE_HISTORY)
    sales    = query(config.SQL_SALE_RESULTS)

    fh_map = defaultdict(list)
    sr_map = defaultdict(list)
    for r in fee_hist: fh_map[str(r.get('stallion_name', ''))].append(r)
    for r in sales:    sr_map[str(r.get('stallion_name', ''))].append(r)

    # Highlights from Excel until a DB table exists
    try:
        hl_map = _load_highlights_from_xl()
    except Exception as e:
        print(f"  WARNING: could not load highlights from Excel ({e})")
        hl_map = defaultdict(list)
    return fh_map, sr_map, hl_map


def load_stallions_from_xl():
    """Load stallion profile data from the Stallions sheet in stallion_data.xlsx.
    Returns a dict keyed by stallion name."""
    if not XL.exists():
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(XL)
    if 'Stallions' not in wb.sheetnames:
        return {}
    rows = list(wb['Stallions'].iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result = {}
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        r = {h: (str(v) if v is not None else '') for h, v in zip(headers, row)}
        name = r.get('name', '')
        if name:
            result[name] = r
    return result


def _load_highlights_from_xl():
    """Load highlights from stallion_data.xlsx. Returns hl_map (defaultdict)."""
    hl_map = defaultdict(list)
    if not XL.exists():
        return hl_map
    import openpyxl
    wb = openpyxl.load_workbook(XL)
    if 'Highlights' not in wb.sheetnames:
        return hl_map
    rows = list(wb['Highlights'].iter_rows(values_only=True))
    if not rows:
        return hl_map
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        r = {h: (v if v is not None else '') for h, v in zip(headers, row)}
        hl_map[str(r.get('stallion_name', ''))].append(r)
    return hl_map


def load_tables_from_xl():
    """Load fee history, sale results, and highlights from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(XL)

    def to_dicts(ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        return [
            {h: (v if v is not None else '') for h, v in zip(headers, row)}
            for row in rows[1:] if not all(v is None for v in row)
        ]

    fee_hist   = to_dicts(wb['FeeHistory'])
    sales      = to_dicts(wb['SaleResults'])
    highlights = to_dicts(wb['Highlights']) if 'Highlights' in wb.sheetnames else []

    fh_map = defaultdict(list)
    sr_map = defaultdict(list)
    hl_map = defaultdict(list)
    for r in fee_hist:   fh_map[str(r.get('stallion_name', ''))].append(r)
    for r in sales:      sr_map[str(r.get('stallion_name', ''))].append(r)
    for r in highlights: hl_map[str(r.get('stallion_name', ''))].append(r)
    return fh_map, sr_map, hl_map


_TDN_SALE_TYPES = [
    ('1', '2YO in Training'),
    ('2', 'Covering'),
    ('3', 'Weanlings'),
    ('4', 'Yearlings'),
]


def _parse_tdn_table(html, year, sale_label, name_map):
    """Parse TDN insta-tistics page HTML for one year + sale type.
    Returns {stallion_name: row_dict} for stallions present in name_map.
    name_map is {normalized_name: original_name}.
    Covering (sale_type=2) has 13 columns with only Top Mare; others have 14
    with Top Colt + Top Filly.
    """
    class _TDNParser(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.in_table    = False
            self.in_tbody    = False
            self.in_tr       = False
            self.in_td       = False
            self.depth       = 0
            self.table_depth = -1
            self.current_cells = []
            self.current_cell  = ''
            self.all_rows      = []

        def handle_starttag(self, tag, attrs):
            self.depth += 1
            d = dict(attrs)
            if tag == 'table' and d.get('id') == 'insta-main':
                self.in_table    = True
                self.table_depth = self.depth
            if not self.in_table:
                return
            if tag == 'tbody':
                self.in_tbody = True
            if self.in_tbody and tag == 'tr':
                self.in_tr         = True
                self.current_cells = []
            if self.in_tr and tag == 'td':
                self.in_td        = True
                self.current_cell = ''

        def handle_endtag(self, tag):
            if self.in_td and tag == 'td':
                self.current_cells.append(self.current_cell.strip())
                self.in_td = False
            if self.in_tr and tag == 'tr':
                if self.current_cells:
                    self.all_rows.append(self.current_cells)
                self.in_tr = False
            if self.in_tbody and tag == 'tbody':
                self.in_tbody = False
            if self.in_table and tag == 'table' and self.depth == self.table_depth:
                self.in_table = False
            self.depth -= 1

        def handle_data(self, data):
            if self.in_td:
                self.current_cell += data

    p = _TDNParser()
    p.feed(html)

    def clean_price(s):
        s = s.strip().replace('$', '').replace(',', '').replace('\xa0', '').strip()
        return '' if s in ('--', '-', '') else s

    def clean_num(s):
        s = s.strip().replace('\xa0', '').strip()
        return '' if s in ('--', '-', '') else s

    is_covering = (sale_label == 'Covering')

    result = {}
    for cols in p.all_rows:
        min_cols = 10 if is_covering else 11
        if len(cols) < min_cols:
            continue
        raw_name = cols[1].strip()
        norm     = re.sub(r'[^a-z0-9]', '', raw_name.lower())
        original = name_map.get(norm)
        if not original:
            continue
        # Covering: col[9] = Top Mare (→ top_filly); no top_colt
        # Others:   col[9] = Top Colt, col[10] = Top Filly
        if is_covering:
            top_colt  = ''
            top_filly = clean_price(cols[9])
        else:
            top_colt  = clean_price(cols[9])
            top_filly = clean_price(cols[10]) if len(cols) > 10 else ''
        result[original] = {
            'stallion_name': original,
            'year':      year,
            'sale_type': sale_label,
            'ring':      clean_num(cols[5]),
            'sold':      clean_num(cols[6]),
            'average':   clean_price(cols[7]),
            'median':    clean_price(cols[8]),
            'top_colt':  top_colt,
            'top_filly': top_filly,
        }
    return result


def scrape_tdn_auction_results(stallion_names):
    """Scrape auction results from TDN insta-tistics for the current and prior year,
    across all four sale types (2YO in Training, Covering, Weanlings, Yearlings).
    Returns sr_map: {stallion_name: [row_dicts]}, newest year first.
    """
    from datetime import date
    current_year = date.today().year
    years = [current_year - 1, current_year]

    name_map = {re.sub(r'[^a-z0-9]', '', n.lower()): n for n in stallion_names}

    sr_map = defaultdict(list)
    TDN_BASE = 'https://www.thoroughbreddailynews.com/insta-tistics/'

    for year in years:
        for sale_code, sale_label in _TDN_SALE_TYPES:
            params = urllib.parse.urlencode({
                'sire': '', 'log': '', 'sortBy': 'sortByYear',
                'txbReportType': '2', 'sale_type': sale_code,
                'selYear': str(year), 'results': '10000',
                'ranked': '1', 'freshmen': '0', 'location': '1',
            })
            url = TDN_BASE + '?' + params
            print(f"  TDN {year} {sale_label}...", end=' ', flush=True)
            try:
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=20) as resp:
                    html = resp.read().decode('utf-8', errors='replace')
            except Exception as e:
                print(f'FAILED ({e})')
                continue
            rows = _parse_tdn_table(html, year, sale_label, name_map)
            for sname, row in rows.items():
                sr_map[sname].append(row)
            print(f'{len(rows)} matched')

    # Sort: newest year first, then alphabetically by sale type within the same year
    _SALE_ORDER = {label: i for i, (_, label) in enumerate(_TDN_SALE_TYPES)}
    for name in sr_map:
        sr_map[name].sort(key=lambda r: (-r['year'], _SALE_ORDER.get(r['sale_type'], 99)))

    return sr_map


def load_tables():
    """Load fee history + highlights from DB/Excel; auction results scraped live from TDN."""
    # Fee history + highlights: try DB, fall back to Excel
    try:
        import config
        print("Loading fee history from SQL database...")
        fh_map, _, hl_map = load_tables_from_db()
        print("  Done.")
    except Exception as e:
        print(f"  DB load failed ({e}), loading from Excel.")
        fh_map, _, hl_map = load_tables_from_xl()

    # Auction results: scrape live from TDN
    print("Scraping auction results from TDN insta-tistics...")
    sr_map = scrape_tdn_auction_results(list(PEDIGREES.keys()))
    if not any(sr_map.values()):
        print("  TDN scrape returned no data, falling back to Excel sale results.")
        _, sr_map, _ = load_tables_from_xl()

    return fh_map, sr_map, hl_map

# ── HTML builders ─────────────────────────────────────────────────────────────

def bold_progeny_name(text):
    """Bold the leading horse name (ALL CAPS sequence) in progeny text."""
    raw = str(text)
    # Match greedy ALL-CAPS name (letters, spaces, apostrophes) + optional parentheticals
    m = re.match(r'^([A-Z][A-Z\' ]+(?:\s*\([^)]+\))*)(\s*[:\-]?\s*)(.*)', raw, re.DOTALL)
    if m and len(m.group(1).strip()) > 1:
        return (f'<strong class="prog-name">{esc(m.group(1))}</strong>'
                f'{esc(m.group(2))}{esc(m.group(3))}')
    return esc(raw)

def pedigree_html(name):
    p = PEDIGREES.get(name)
    if not p:
        return ''
    def pc(k): return esc(p.get(k, '—'))
    return f'''
      <div class="content-block ped-block">
        <h3 class="block-title"><span class="block-icon">🧬</span>Pedigree<span class="chevron">▾</span></h3>
        <div class="block-body">
          <div class="ped-grid">
            <div class="ped-col ped-col-parent">
              <div class="ped-cell ped-sire-cell">{pc("sire")}</div>
              <div class="ped-cell ped-dam-cell">{pc("dam")}</div>
            </div>
            <div class="ped-col ped-col-grand">
              <div class="ped-cell">{pc("sire_sire")}</div>
              <div class="ped-cell ped-dam-side">{pc("sire_dam")}</div>
              <div class="ped-cell">{pc("dam_sire")}</div>
              <div class="ped-cell ped-dam-side">{pc("dam_dam")}</div>
            </div>
          </div>
        </div>
      </div>'''

def pedigree_highlights_html(highlights):
    """Return a Pedigree Highlights block for items with category='pedigree_highlight', or ''."""
    items = [h for h in highlights if str(h.get('category', '')).strip() == 'pedigree_highlight']
    if not items:
        return ''
    lis = ''.join(f'<li>{esc(str(h.get("text", "")))}</li>' for h in items)
    return f'''
      <div class="content-block">
        <h3 class="block-title"><span class="block-icon">🧬</span>Pedigree Highlights<span class="chevron">▾</span></h3>
        <div class="block-body">
          <ul class="bullet-list">{lis}</ul>
        </div>
      </div>'''

def selling_points_html(name, points):
    if not points:
        return ''
    lis = ''.join(f'<li>{p}</li>' for p in points)
    return f'''
      <div class="content-block sp-block">
        <h3 class="block-title"><span class="block-icon">⭐</span>Why {esc(name)}<span class="chevron">▾</span></h3>
        <div class="block-body">
          <ul class="sp-list">{lis}</ul>
        </div>
      </div>'''

def more_selling_points_html(highlights):
    """Return a More Selling Points block for items with category='more_selling_point', or ''."""
    items = [h for h in highlights if str(h.get('category', '')).strip() == 'more_selling_point']
    if not items:
        return ''
    lis = ''.join(f'<li>{esc(str(h.get("text", "")))}</li>' for h in items)
    return f'''
      <div class="content-block sp-block">
        <h3 class="block-title"><span class="block-icon">🔍</span>More Selling Points<span class="chevron">▾</span></h3>
        <div class="block-body">
          <ul class="sp-list">{lis}</ul>
        </div>
      </div>'''

def _crop_age_label(season):
    """Return e.g. '3yo', 'yearlings', 'weanlings' for a given breeding season."""
    from datetime import date
    try:
        age = date.today().year - int(season) - 1
    except (ValueError, TypeError):
        return ''
    if age < 0:
        return ''
    if age == 0:
        return 'foals'
    if age == 1:
        return 'yearlings'
    return f'{age}yo'

def fee_table_html(fee_hist):
    if not fee_hist:
        return ''
    rows = ''
    for fh in fee_hist:
        season  = fh.get('season', '')
        foals   = fh.get('Foals', fh.get('foals', ''))
        foals   = '' if foals is None or str(foals).strip().lower() == 'none' else foals
        age_lbl = _crop_age_label(season)
        foals_cell = esc(str(foals)) if foals != '' else ''
        if foals_cell and age_lbl:
            foals_cell += f' <span class="crop-age">{age_lbl}</span>'
        def nv(val):
            return '' if val is None or str(val).strip().lower() == 'none' else val

        rows += f'''
          <tr>
            <td class="bold">{esc(nv(season))}</td>
            <td class="fee-cell">{esc(fmt_currency(nv(fh.get('stud_fee',''))))}</td>
            <td class="center">{esc(nv(fh.get('mares_bred','')))}</td>
            <td class="center">{esc(nv(fh.get('CI', fh.get('ci',''))))}</td>
            <td class="center">{esc(nv(fh.get('CPI', fh.get('cpi',''))))}</td>
            <td>{foals_cell}</td>
            <td class="center">{esc(nv(str(fh.get('runners', '') or '')))}</td>
            <td class="center">{esc(nv(str(fh.get('black_type_winners', '') or '')))}</td>
            <td class="center">{esc(_fmt_pct(fh.get('SW_percent', '')))}</td>
          </tr>'''
    notes = '  ·  '.join(str(fh.get('notes','')) for fh in fee_hist if fh.get('notes'))
    notes_html = f'<p class="table-note">{esc(notes)}</p>' if notes else ''
    return f'''
      <div class="content-block" id="fees">
        <h3 class="block-title"><span class="block-icon">💰</span>Stud Fee History<span class="chevron">▾</span></h3>
        <div class="block-body">
          <div class="table-scroll-wrap"><div class="table-scroll">
            <table>
              <thead><tr>
                <th>Season</th><th>Stud Fee</th><th class="center">Mares Bred</th>
                <th class="center">CI</th><th class="center">CPI</th><th>Foals</th><th class="center">Runners</th><th class="center">Stakes Winners</th><th class="center">%SW</th>
              </tr></thead>
              <tbody>{rows}</tbody>
            </table>
          </div></div>
          {notes_html}
        </div>
      </div>'''

def sales_table_html(sales):
    if not sales:
        return ''
    rows = ''
    def sv(val):
        """Return empty string for None/null DB values."""
        return '' if val is None or str(val).strip().lower() == 'none' else val

    for sr in sales:
        rows += f'''
          <tr>
            <td class="bold">{esc(sv(sr.get('year','')))}</td>
            <td>{esc(sv(sr.get('sale_type','')))}</td>
            <td class="center">{esc(sv(sr.get('ring','')))}</td>
            <td class="center">{esc(sv(sr.get('sold','')))}</td>
            <td class="bold center">{esc(fmt_currency(sv(sr.get('average',''))))}</td>
            <td class="center">{esc(fmt_currency(sv(sr.get('median',''))))}</td>
            <td class="center">{esc(fmt_currency(sv(sr.get('top_colt',''))))}</td>
            <td class="center">{esc(fmt_currency(sv(sr.get('top_filly',''))))}</td>
          </tr>'''
    return f'''
      <div class="content-block">
        <h3 class="block-title"><span class="block-icon">🔨</span>Auction Results<span class="chevron">▾</span></h3>
        <div class="block-body">
          <div class="table-scroll-wrap"><div class="table-scroll">
            <table>
              <thead><tr>
                <th>Year</th><th>Sale Type</th><th class="center">Ring</th><th class="center">Sold</th>
                <th class="center">Average</th><th class="center">Median</th><th class="center">Top Colt</th><th class="center">Top Filly / Mare</th>
              </tr></thead>
              <tbody>{rows}</tbody>
            </table>
          </div></div>
        </div>
      </div>'''

def race_record_html(name):
    """Return a Race Record block with the PP image, or '' if no file exists."""
    slug     = _stallion_slug(name)
    pp_file  = PP_DIR / f"{slug}.png"
    if not pp_file.exists():
        return ''
    # Use a relative path from the HTML output file
    rel_path = f"stallion PPs png files/{slug}.png"
    encoded  = rel_path.replace(' ', '%20')
    return f'''
      <div class="content-block pp-block">
        <h3 class="block-title"><span class="block-icon">📋</span>Race Record<span class="chevron">▾</span></h3>
        <div class="block-body">
          <div class="pp-wrap">
            <img src="{encoded}" alt="{esc(name)} past performances" class="pp-img" loading="lazy">
          </div>
        </div>
      </div>'''


def highlights_html(highlights):
    gen  = [h for h in highlights if str(h.get('category','')).strip() == 'general']
    sp   = [h for h in highlights if str(h.get('category','')).strip() == 'selling_point']
    prog = [h for h in highlights if str(h.get('category','')).strip() == 'progeny']
    ages = {
        'at_two':   [h for h in highlights if str(h.get('category','')).strip() == 'at_two'],
        'at_three': [h for h in highlights if str(h.get('category','')).strip() == 'at_three'],
        'at_four':  [h for h in highlights if str(h.get('category','')).strip() == 'at_four'],
        'at_five':  [h for h in highlights if str(h.get('category','')).strip() == 'at_five'],
    }

    has_career = gen or sp or any(ages.values())
    out = ''

    if has_career:
        items_html = ''
        if sp:
            items_html += '<ul class="bullet-list">'
            for h in sp:
                items_html += f'<li>{esc(str(h.get("text","")))}</li>'
            items_html += '</ul>'
        if gen:
            items_html += '<ul class="bullet-list">'
            for h in gen:
                items_html += f'<li>{esc(str(h.get("text","")))}</li>'
            items_html += '</ul>'
        for cat, label in [('at_two','At two'),('at_three','At three'),
                            ('at_four','At four'),('at_five','At five')]:
            items = ages.get(cat, [])
            if not items:
                continue
            subtitle = str(items[0].get('subtitle', ''))
            sub_html = (f'<span class="age-subtitle">{esc(subtitle)}</span>'
                        if subtitle else '')
            lis = ''.join(f'<li>{esc(str(h.get("text","")))}</li>' for h in items)
            items_html += f'''
              <div class="age-group">
                <div class="age-label">{esc(label)}{sub_html}</div>
                <ul class="bullet-list">{lis}</ul>
              </div>'''
        out += f'''
      <div class="content-block highlights-block">
        <h3 class="block-title"><span class="block-icon">🏆</span>Career Highlights<span class="chevron">▾</span></h3>
        <div class="block-body">
          <div class="highlights">{items_html}</div>
        </div>
      </div>'''

    if prog:
        lis = ''.join(f'<li>{bold_progeny_name(h.get("text",""))}</li>' for h in prog)
        out += f'''
      <div class="content-block">
        <h3 class="block-title"><span class="block-icon">🐎</span>Notable Progeny<span class="chevron">▾</span></h3>
        <div class="block-body">
          <ul class="progeny-list">{lis}</ul>
        </div>
      </div>'''

    return out

def stallion_section(s, fee_hist, sales, highlights, sp_points):
    name      = str(s.get('name', ''))
    slug      = slugify(name)
    fee       = str(s.get('fee_current', ''))
    sire      = str(s.get('sire', ''))
    dam       = str(s.get('dam', ''))
    damsire   = str(s.get('damsire', ''))
    foaled    = str(s.get('year_foaled', ''))
    height    = str(s.get('height', ''))
    earnings  = str(s.get('earnings', ''))
    entered   = str(s.get('entered_stud', ''))
    fcrop     = str(s.get('first_crop_note', ''))
    photo_f   = str(s.get('photo_file',
                          name.lower().replace(' ','_').replace("'","") + '.jpg'))
    photo_src = PHOTO_URLS.get(photo_f, f'img_cache/{photo_f}')

    name_parts = name.upper().split()
    name_html  = ' '.join(
        f'<span class="name-word">{esc(w)}</span>' for w in name_parts
    )

    info_items = []
    if foaled:   info_items.append(esc(foaled))
    if height:   info_items.append(esc(height))
    if earnings: info_items.append(f'Earnings: {esc(earnings)}')
    if entered:  info_items.append(f'Entered stud: {esc(entered)}')
    if fcrop:    info_items.append(esc(fcrop))
    info_html  = ''.join(f'<span class="info-pill">{i}</span>' for i in info_items)
    fcrop_html = ''
    accord_fee = f'<span class="accord-fee">{esc(fee)}</span>' if fee else ''
    return f'''
  <section class="stallion" id="{slug}">
    <button class="accord-header" aria-expanded="false">
      <img class="accord-thumb" src="{esc(photo_src)}" alt="" aria-hidden="true">
      <span class="accord-name">{esc(name.upper())}</span>
      {accord_fee}
      <span class="accord-chev">▾</span>
    </button>
    <div class="sticky-bar">
      <span class="sb-name">{esc(name.upper())}</span>
      {'<span class="sb-fee">' + esc(fee) + '</span>' if fee else ''}
    </div>
    <div class="stallion-card">

      <!-- ── HEADER ── -->
      <div class="stallion-header">
        <div class="photo-col">
          <img src="{esc(photo_src)}" alt="{esc(name)}" loading="lazy">
        </div>
        <div class="info-col">
          <div class="info-title">
            <span class="info-name">{esc(name.upper())}</span>
            <button class="share-btn" data-slug="{slug}" title="Copy link">🔗</button>
          </div>
          <div class="info-fee">{esc(fee)}</div>
          <div class="ped-line">{esc(sire)} <span class="ped-dash">—</span> {esc(dam)} <span class="ped-paren">({esc(damsire)})</span></div>
          <div class="info-pills">{info_html}</div>
          {fcrop_html}
        </div>
      </div>

      <!-- ── BODY ── -->
      <div class="stallion-body">
        {selling_points_html(name, sp_points)}
        {more_selling_points_html(highlights)}
        {fee_table_html(fee_hist)}
        {highlights_html(highlights)}
        {sales_table_html(sales)}
        {race_record_html(name)}
        {pedigree_html(name)}
        {pedigree_highlights_html(highlights)}
        <div class="card-foot">
          <a href="#{slug}" class="back-to-card">↑ Back to top of card</a>
          <button class="print-btn" onclick="window.print()">⎙ Print</button>
        </div>
      </div>

    </div>
  </section>'''

# ── CSS ───────────────────────────────────────────────────────────────────────

CSS = """
/* ── Reset & base ── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html { scroll-behavior: smooth; font-size: 15px; }
body {
  font-family: 'FF Meta', 'Meta', 'Source Sans 3', 'Segoe UI', system-ui, Arial, sans-serif;
  color: #222;
  background: #e8eef6;
  line-height: 1.5;
}

/* ── Brand vars ── */
:root {
  --blue:   #0037B2;
  --blue2:  #002A85;
  --cyan:   #00ABEE;
  --cyan-lt:#E3F5FD;
  --red:    #E3140D;
  --grey:   #5a6a7a;
  --border: #c8d8ec;
  --light:  #f0f6fc;
  --white:  #ffffff;
  --card-shadow: 0 2px 20px rgba(0, 55, 178, 0.13);
  --radius: 10px;
}

/* ── Navbar ── */
#navbar {
  position: sticky;
  top: 0;
  z-index: 200;
  background: #0137B3;
  display: flex;
  align-items: center;
  gap: 12px;
  padding: 0 16px;
  height: 54px;
  box-shadow: 0 3px 10px rgba(0,0,0,0.35);
}
.nav-logo { flex-shrink: 0; height: 34px; display: flex; align-items: center; }
.nav-logo img { height: 100%; width: auto; object-fit: contain; }
.nav-divider { width: 1px; height: 28px; background: rgba(255,255,255,0.2); flex-shrink: 0; }
.nav-links {
  display: flex;
  gap: 5px;
  overflow-x: auto;
  padding: 6px 0;
  flex: 1;
  scrollbar-width: thin;
  scrollbar-color: var(--cyan) transparent;
}
.nav-links::-webkit-scrollbar { height: 3px; }
.nav-links::-webkit-scrollbar-thumb { background: var(--cyan); border-radius: 2px; }
.nav-links a {
  color: rgba(255,255,255,0.80);
  text-decoration: none;
  font-size: 11.5px;
  font-weight: 600;
  white-space: nowrap;
  padding: 5px 11px;
  border-radius: 20px;
  border: 1px solid rgba(255,255,255,0.18);
  transition: background 0.18s, color 0.18s, border-color 0.18s;
  letter-spacing: 0.2px;
}
.nav-links a:hover { background: rgba(0,171,238,0.25); color: white; border-color: rgba(0,171,238,0.5); }
.nav-links a.active { background: var(--cyan); color: white; border-color: var(--cyan); }

/* ── Hero ── */
.hero {
  background: linear-gradient(160deg, #001f6e 0%, var(--blue) 60%, #0057c8 100%);
  color: white;
  text-align: center;
  padding: 52px 20px 40px;
  position: relative;
  overflow: hidden;
}
.hero::after {
  content: '';
  position: absolute;
  bottom: 0; left: 0; right: 0;
  height: 4px;
  background: #cc1012;
}
.hero-updated {
  display: inline-block;
  margin-left: 8px;
  background: #cc1012;
  color: white;
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1px;
  padding: 2px 7px;
  border-radius: 3px;
  vertical-align: middle;
  text-transform: uppercase;
}
.hero h1 {
  font-family: 'Egizio', 'Playfair Display', Georgia, serif;
  font-size: clamp(22px, 4.5vw, 44px);
  font-weight: 800;
  letter-spacing: 4px;
  margin-bottom: 8px;
}
.hero p {
  color: rgba(255,255,255,0.6);
  font-size: clamp(11px, 1.5vw, 13px);
  letter-spacing: 2px;
  text-transform: uppercase;
}
.hero .stallion-count {
  display: inline-block;
  margin-top: 18px;
  background: rgba(0,171,238,0.2);
  border: 1px solid rgba(0,171,238,0.4);
  color: rgba(255,255,255,0.85);
  font-size: 12px;
  padding: 5px 16px;
  border-radius: 20px;
  letter-spacing: 1px;
}

/* ── Main container ── */
main { max-width: 1180px; margin: 0 auto; padding: 28px 16px 48px; }

/* Accordion header — hidden on desktop, shown on mobile */
.accord-header { display: none; }

/* ── Stallion section ── */
.stallion { scroll-margin-top: 62px; margin-bottom: 32px; }
.stallion-card {
  background: var(--white);
  border-radius: var(--radius);
  overflow: hidden;
  box-shadow: var(--card-shadow);
  border: 1px solid var(--border);
}

/* ── Stallion header ── */
.stallion-header {
  display: grid;
  grid-template-columns: 300px 1fr;
  border-bottom: 1px solid var(--border);
  background: var(--light);
}

/* Name column */
.name-col {
  background: var(--blue2);
  padding: 20px 14px 16px;
  display: flex;
  flex-direction: column;
  align-items: flex-start;
  justify-content: space-between;
  border-right: 3px solid var(--cyan);
}
.big-name {
  display: flex;
  flex-direction: column;
  gap: 2px;
}
.name-word {
  display: block;
  font-family: 'Egizio', 'Playfair Display', Georgia, serif;
  font-size: clamp(18px, 1.6vw, 24px);
  font-weight: 900;
  color: var(--cyan);
  line-height: 1.05;
  letter-spacing: 0.5px;
  white-space: nowrap;
}
.big-fee {
  font-size: 13px;
  font-weight: 700;
  color: rgba(255,255,255,0.75);
  margin-top: 8px;
  letter-spacing: 0.3px;
}

/* Photo column */
.photo-col {
  overflow: hidden;
  aspect-ratio: 4 / 3;
  border-right: 1px solid var(--border);
  background: white;
  display: flex;
  align-items: center;
  justify-content: center;
}
.photo-col img {
  width: 100%;
  height: 100%;
  object-fit: contain;
  display: block;
}

/* Info column */
.info-col {
  padding: 18px 20px 16px;
  display: flex;
  flex-direction: column;
  justify-content: center;
  gap: 7px;
}
.info-title {
  display: flex;
  flex-wrap: wrap;
  align-items: baseline;
  gap: 10px;
  margin-bottom: 2px;
}
.info-name {
  font-family: 'Egizio', 'Playfair Display', Georgia, serif;
  font-size: clamp(26px, 3vw, 38px);
  font-weight: 900;
  color: var(--blue);
  letter-spacing: 1px;
  line-height: 1.05;
  text-decoration: underline;
  text-decoration-color: var(--cyan);
  text-underline-offset: 5px;
  text-decoration-thickness: 3px;
}
.info-fee {
  font-size: clamp(14px, 1.5vw, 17px);
  font-weight: 700;
  color: #cc1012;
}
.ped-line {
  font-size: 14px;
  font-weight: 600;
  color: var(--blue);
}
.ped-dash { color: var(--cyan); margin: 0 4px; }
.ped-paren { color: var(--grey); font-weight: 400; }
.info-pills { display: flex; flex-wrap: wrap; gap: 5px; margin-top: 2px; }
.info-pill {
  font-size: 11.5px;
  color: var(--grey);
  background: white;
  border: 1px solid var(--border);
  border-radius: 12px;
  padding: 2px 10px;
}
.first-crop {
  font-size: 12px;
  color: var(--cyan);
  font-style: italic;
  margin-top: 2px;
}

/* ── Stallion body ── */
.stallion-body { }

/* Content block */
.content-block {
  padding: 18px 22px;
  border-top: 1px solid var(--border);
}
.content-block:first-child { border-top: none; }

.block-title {
  font-size: 10.5px;
  font-weight: 700;
  letter-spacing: 1.8px;
  text-transform: uppercase;
  color: var(--blue);
  margin-bottom: 13px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.block-title::before {
  content: '';
  display: inline-block;
  width: 3px;
  height: 14px;
  background: var(--cyan);
  border-radius: 2px;
  flex-shrink: 0;
}
.block-icon {
  font-size: 14px;
  line-height: 1;
  flex-shrink: 0;
}

/* ── Tables ── */
.table-scroll { overflow-x: auto; border-radius: 6px; border: 1px solid var(--border); }
table { width: 100%; border-collapse: collapse; font-size: 12.5px; min-width: 560px; }
thead th {
  background: var(--blue);
  color: white;
  padding: 9px 11px;
  text-align: left;
  font-size: 11px;
  font-weight: 600;
  letter-spacing: 0.4px;
  white-space: nowrap;
}
thead th:first-child { border-radius: 6px 0 0 0; }
thead th:last-child  { border-radius: 0 6px 0 0; }
tbody tr { transition: background 0.12s; }
tbody tr:nth-child(even)  { background: var(--cyan-lt); }
tbody tr:nth-child(odd)   { background: white; }
tbody tr:hover { background: #d0ecf9; }
td {
  padding: 7px 11px;
  border-bottom: 1px solid var(--border);
  vertical-align: top;
  color: #333;
  line-height: 1.35;
}
tbody tr:last-child td { border-bottom: none; }
td.bold   { font-weight: 700; }
td.center { text-align: center; }
td.money  { text-align: right; font-variant-numeric: tabular-nums; }
td.small  { font-size: 11.5px; }
td.fee-cell { font-weight: 700; color: var(--cyan); }
.table-note { font-size: 11.5px; color: var(--grey); font-style: italic; margin-top: 7px; }
th.center { text-align: center; }
.crop-age { font-size: 10px; color: var(--grey); font-weight: 400; margin-left: 3px; }

/* ── Selling points ── */
.sp-block { background: linear-gradient(135deg, #f0f6ff 0%, #e8f4fd 100%); border-top: 3px solid #cc1012 !important; }
.sp-list { list-style: none; display: flex; flex-direction: column; gap: 6px; }
.sp-list li {
  font-size: 13.5px;
  line-height: 1.5;
  color: #1a2a4a;
  padding: 8px 14px 8px 36px;
  position: relative;
  border-radius: 4px;
}
.sp-list li::before {
  content: '✦';
  position: absolute;
  left: 12px;
  top: 8px;
  color: var(--cyan);
  font-size: 11px;
}

/* ── Highlights ── */
.highlights { display: flex; flex-direction: column; gap: 10px; }
.age-group  { }
.age-label {
  font-size: 13px;
  font-weight: 700;
  color: var(--blue);
  margin-bottom: 5px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.age-subtitle {
  color: var(--cyan);
  font-size: 12px;
  font-weight: 600;
}
.bullet-list { list-style: none; padding-left: 4px; display: flex; flex-direction: column; gap: 4px; }
.bullet-list li {
  font-size: 13px;
  color: #333;
  padding-left: 18px;
  position: relative;
  line-height: 1.45;
}
.bullet-list li::before {
  content: '';
  position: absolute;
  left: 4px;
  top: 7px;
  width: 6px;
  height: 6px;
  background: var(--cyan);
  border-radius: 50%;
}

/* ── Progeny list ── */
.progeny-list { list-style: none; display: flex; flex-direction: column; gap: 6px; }
.progeny-list li {
  font-size: 13px;
  line-height: 1.45;
  padding: 9px 14px 9px 16px;
  background: var(--light);
  border-radius: 5px;
  border-left: 3px solid var(--cyan);
}
.prog-name { font-weight: 700; color: var(--blue); }

/* ── Back-to-top ── */
.back-to-top {
  position: fixed;
  bottom: 24px;
  right: 24px;
  width: 42px;
  height: 42px;
  background: var(--blue);
  color: white;
  border-radius: 50%;
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 18px;
  text-decoration: none;
  box-shadow: 0 3px 12px rgba(0,0,0,0.3);
  opacity: 0;
  transform: translateY(10px);
  transition: opacity 0.25s, transform 0.25s, background 0.2s;
  pointer-events: none;
  z-index: 100;
}
.back-to-top.visible { opacity: 1; transform: none; pointer-events: auto; }
.back-to-top:hover { background: var(--cyan); }

/* ── Footer ── */
footer {
  background: var(--blue2);
  color: rgba(255,255,255,0.55);
  text-align: center;
  padding: 18px 20px;
  font-size: 11px;
  letter-spacing: 0.8px;
  border-top: 3px solid var(--cyan);
}

/* ── Fee callout ── */
.fee-callout {
  display: inline-flex;
  flex-direction: column;
  background: var(--blue);
  border-left: 3px solid var(--cyan);
  border-radius: 6px;
  padding: 5px 14px 6px;
}
.fee-label { font-size: 8.5px; letter-spacing: 1.5px; text-transform: uppercase; color: rgba(255,255,255,0.6); font-weight: 600; }
.fee-amount { font-size: 22px; font-weight: 800; color: var(--cyan); font-family: 'Egizio','Playfair Display',Georgia,serif; line-height: 1.15; }

/* ── Search bar ── */
.search-wrap {
  background: var(--blue2);
  padding: 0 16px 14px;
  display: flex;
  justify-content: center;
}
.search-wrap input {
  width: 100%; max-width: 480px;
  padding: 10px 18px 10px 42px;
  border-radius: 24px;
  border: 2px solid rgba(255,255,255,0.15);
  background: rgba(255,255,255,0.1) url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='rgba(255,255,255,0.5)' stroke-width='2.5'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cpath d='m21 21-4.35-4.35'/%3E%3C/svg%3E") no-repeat 14px center;
  color: white; font-size: 14px; font-family: inherit; outline: none;
  transition: border-color 0.2s, background-color 0.2s;
}
.search-wrap input::placeholder { color: rgba(255,255,255,0.4); }
.search-wrap input:focus { border-color: var(--cyan); background-color: rgba(255,255,255,0.15); }
.search-no-results { text-align: center; padding: 40px 20px; color: var(--grey); font-size: 15px; display: none; }

/* ── Collapsible sections ── */
.block-title { cursor: pointer; user-select: none; }
.chevron { margin-left: auto; font-size: 12px; color: var(--grey); transition: transform 0.25s; flex-shrink: 0; }
.content-block.collapsed .chevron { transform: rotate(-90deg); }
.block-body { overflow: hidden; transition: max-height 0.32s ease, opacity 0.25s ease; }
.content-block.collapsed .block-body { opacity: 0; }

/* ── Pedigree ── */
.ped-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 0; border: 1px solid var(--border); border-radius: 6px; overflow: hidden; }
.ped-col { display: flex; flex-direction: column; }
.ped-col-parent .ped-cell { flex: 1; display: flex; align-items: center; padding: 10px 14px; font-weight: 700; font-size: 14px; color: var(--blue); border-bottom: 1px solid var(--border); }
.ped-col-parent .ped-cell:last-child { border-bottom: none; }
.ped-col-grand .ped-cell { flex: 1; padding: 7px 12px; font-size: 12.5px; color: #333; border-left: 1px solid var(--border); border-bottom: 1px solid var(--border); background: white; }
.ped-col-grand .ped-cell:last-child { border-bottom: none; }
.ped-sire-cell { background: var(--cyan-lt); }
.ped-dam-cell  { background: #f5f0fc; }
.ped-dam-side  { background: #faf8fe; }

/* ── Share button ── */
.share-btn { background: none; border: 1px solid var(--border); border-radius: 6px; padding: 3px 9px; cursor: pointer; font-size: 13px; color: var(--grey); transition: all 0.15s; flex-shrink: 0; line-height: 1.4; }
.share-btn:hover { background: var(--cyan); border-color: var(--cyan); color: white; }

/* ── Race Record (PP image) ── */
.pp-wrap { overflow-x: auto; }
.pp-img  { display: block; max-width: 100%; height: auto; border: 1px solid var(--border); border-radius: 4px; }

/* ── Back to top of card ── */
.card-foot { padding: 10px 22px 16px; display: flex; justify-content: center; border-top: 1px solid var(--border); background: var(--light); }
.back-to-card { font-size: 11.5px; color: var(--grey); text-decoration: none; padding: 5px 18px; border: 1px solid var(--border); border-radius: 20px; transition: all 0.15s; background: white; }
.back-to-card:hover { color: var(--blue); border-color: var(--cyan); background: var(--cyan-lt); }

/* ── Toast ── */
.toast { position: fixed; bottom: 80px; left: 50%; transform: translateX(-50%) translateY(8px); background: var(--blue); color: white; padding: 9px 22px; border-radius: 20px; font-size: 13px; font-weight: 600; opacity: 0; transition: opacity 0.2s, transform 0.2s; pointer-events: none; z-index: 300; border: 1px solid rgba(0,171,238,0.4); box-shadow: 0 4px 16px rgba(0,0,0,0.25); white-space: nowrap; }
.toast.show { opacity: 1; transform: translateX(-50%) translateY(0); }

/* ── Card entrance animation (applied by JS) ── */
.card-anim { opacity: 0; transform: translateY(22px); transition: opacity 0.45s ease, transform 0.45s ease; }
.card-anim.visible { opacity: 1; transform: none; }

/* ── Responsive — tablet ── */
@media (max-width: 860px) {
  .stallion-header {
    grid-template-columns: 1fr;
    grid-template-rows: auto auto;
  }
  .photo-col { grid-row: 1; grid-column: 1; aspect-ratio: 4 / 3; border-right: none; }
  .info-col  { grid-row: 2; grid-column: 1; border-top: 3px solid var(--cyan); }
}

/* ── Responsive — mobile ── */
@media (max-width: 600px) {

  /* Navbar — logo only, hide stallion links */
  #navbar { height: 52px; gap: 8px; padding: 0 12px; }
  .nav-logo img { height: 28px; }
  .nav-divider, .nav-links { display: none; }

  /* Hero */
  .hero { padding: 22px 16px 18px; }
  .hero h1 { letter-spacing: 2px; }
  .stallion-count { font-size: 11px; margin-top: 10px; }

  /* Sticky bar — hidden by default on mobile, shown only when accordion is open and scrolled */
  .sticky-bar { display: none !important; top: 52px; padding: 6px 14px; }
  .sticky-bar.sb-visible { display: flex !important; }

  /* Accordion list replaces card layout */
  main { padding: 0 0 40px; max-width: 100%; background: var(--blue2); }

  .stallion { scroll-margin-top: 56px; margin-bottom: 0; border-bottom: 1px solid rgba(255,255,255,0.1); }

  /* Accordion header — full-width tap target */
  .accord-header {
    display: flex;
    align-items: center;
    width: 100%;
    padding: 9px 16px;
    background: var(--blue2);
    border: none;
    color: white;
    cursor: pointer;
    text-align: left;
    gap: 10px;
    -webkit-tap-highlight-color: rgba(0,171,238,0.15);
  }
  .accord-header:active { opacity: 0.85; }
  .stallion.accord-open .accord-header { border-bottom: 2px solid var(--cyan); }
  .accord-thumb {
    width: 46px;
    height: 34px;
    object-fit: cover;
    object-position: center top;
    border-radius: 3px;
    flex-shrink: 0;
    background: white;
  }
  .accord-name {
    font-family: 'Egizio', 'Playfair Display', Georgia, serif;
    font-size: 18px;
    font-weight: 900;
    letter-spacing: 0.5px;
    flex: 1;
  }
  .accord-fee { font-size: 13px; font-weight: 700; color: #cc1012; white-space: nowrap; }
  /* Alternating row colors */
  .stallion:nth-child(odd)  .accord-header { background: var(--blue2); }
  .stallion:nth-child(even) .accord-header { background: var(--cyan); color: var(--blue2); }
  .stallion:nth-child(even) .accord-name   { color: var(--blue2); }
  .stallion:nth-child(even) .accord-fee    { color: #cc1012; }
  .stallion:nth-child(even) .accord-chev   { color: var(--blue2); opacity: 0.6; }
  .stallion:nth-child(even).accord-open .accord-chev { opacity: 1; color: var(--blue2); }
  .accord-chev { font-size: 16px; color: rgba(255,255,255,0.5); transition: transform 0.25s; flex-shrink: 0; }
  .stallion.accord-open .accord-chev { transform: rotate(180deg); color: var(--cyan); }

  /* Card hidden by default on mobile, revealed when open */
  .stallion-card { display: none; border-radius: 0; box-shadow: none; border: none; border-top: none; }
  .stallion.accord-open .stallion-card { display: block; }

  /* Inside the open card */
  .stallion-header { grid-template-columns: 1fr; }
  .photo-col { aspect-ratio: 16 / 10; border-right: none; }
  .info-col { border-top: 3px solid var(--cyan); padding: 14px 16px 12px; gap: 5px; }
  .info-name { font-size: 22px; }
  .info-fee  { font-size: 13px; }
  .ped-line  { font-size: 13px; }
  .info-pill { font-size: 11px; padding: 2px 8px; }

  /* Content blocks */
  .content-block { padding: 14px 12px; }
  .block-title { font-size: 9.5px; letter-spacing: 1.4px; margin-bottom: 10px; }

  /* Selling points */
  .sp-list li { font-size: 13px; padding: 7px 10px 7px 30px; }
  .sp-list li::before { left: 10px; }

  /* Highlights */
  .bullet-list li { font-size: 12.5px; }
  .age-label { font-size: 12.5px; }

  /* Progeny */
  .progeny-list li { font-size: 12.5px; padding: 8px 10px 8px 12px; }

  /* Tables */
  .table-scroll { border-radius: 4px; }
  table { font-size: 11px; min-width: 500px; }
  thead th { padding: 7px 8px; font-size: 10px; }
  td { padding: 5px 8px; }

  /* Back-to-top */
  .back-to-top { bottom: 16px; right: 14px; width: 38px; height: 38px; font-size: 16px; }

  /* Footer */
  footer { font-size: 10px; padding: 14px 16px; letter-spacing: 0.4px; }
}

/* ── Sticky name bar (desktop) ── */
.sticky-bar {
  display: none;
  position: sticky;
  top: 54px;
  z-index: 50;
  background: white;
  border-bottom: 3px solid var(--cyan);
  padding: 7px 22px;
  align-items: center;
  gap: 14px;
  box-shadow: 0 2px 8px rgba(0,55,178,0.10);
}
.sticky-bar.sb-visible { display: flex; }
.sticky-bar .sb-name {
  font-family: 'Egizio', 'Playfair Display', Georgia, serif;
  font-size: 15px;
  font-weight: 900;
  color: var(--blue);
  letter-spacing: 1px;
  text-transform: uppercase;
  flex: 1;
}
.sticky-bar .sb-fee { font-size: 12px; font-weight: 700; color: #cc1012; }

/* ── Print button ── */
.print-btn {
  font-family: inherit; font-size: 11.5px; color: var(--grey); padding: 5px 18px;
  border: 1px solid var(--border); border-radius: 20px;
  background: white; cursor: pointer; transition: all 0.15s;
}
.print-btn:hover { color: var(--blue); border-color: var(--cyan); background: var(--cyan-lt); }

@page { size: letter; margin: 0.35in; }

@media print {
  * { -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }

  #navbar, .hero, .accord-header, .share-btn, .back-to-top, .card-foot,
  footer, .sticky-bar, .print-btn { display: none !important; }

  body, main { background: white !important; padding: 0 !important; margin: 0 !important; }
  main { max-width: 100% !important; }

  /* One stallion per page */
  .stallion { page-break-after: always; break-after: page; margin: 0 !important; }
  .stallion:last-child { page-break-after: avoid; break-after: avoid; }
  .stallion-card { display: block !important; box-shadow: none !important;
    border: 1px solid #ccc !important; border-radius: 0 !important; overflow: visible !important;
    opacity: 1 !important; transform: none !important; }

  /* Compact header */
  .stallion-header { grid-template-columns: 150px 1fr !important; min-height: 0 !important; }
  .photo-col { aspect-ratio: 4/3 !important; }
  .info-col { padding: 10px 14px !important; gap: 4px !important; }
  .info-name { font-size: 18px !important; }
  .info-fee  { font-size: 12px !important; }
  .ped-line  { font-size: 11px !important; }
  .info-pill { font-size: 10px !important; padding: 2px 6px !important; }
  .first-crop { font-size: 10px !important; }

  /* Two-column body sections */
  .stallion-body { display: grid !important; grid-template-columns: 1fr 1fr !important;
    gap: 0 !important; align-items: start !important; }

  /* Expand all collapsed sections */
  .block-body { max-height: none !important; overflow: visible !important;
    display: block !important; padding-bottom: 4px !important; }
  .chevron { display: none !important; }

  /* Compact section blocks */
  .content-block { break-inside: avoid; page-break-inside: avoid;
    padding: 7px 10px !important; border-top: 1px solid #ddd !important; }
  .block-title { font-size: 7.5px !important; letter-spacing: 1px !important;
    margin-bottom: 5px !important; padding-bottom: 4px !important; }

  /* Selling points */
  .sp-list li { font-size: 10px !important; padding: 2px 8px 2px 20px !important; line-height: 1.4 !important; }

  /* Highlights / progeny */
  .bullet-list li, .age-label { font-size: 10px !important; padding: 1px 0 !important; }
  .progeny-list li { font-size: 10px !important; padding: 4px 8px 4px 10px !important; }

  /* Tables */
  .table-scroll { overflow: visible !important; }
  table { font-size: 9.5px !important; min-width: 0 !important; width: 100% !important; }
  thead th { padding: 4px 6px !important; font-size: 8.5px !important; }
  td { padding: 3px 6px !important; }

  /* Pedigree */
  .ped-grid { font-size: 10px !important; gap: 4px !important; }
  .ped-cell { padding: 3px 6px !important; }
}
"""

# ── JavaScript ────────────────────────────────────────────────────────────────

JS = """
(function() {

  // ── Active nav link ──
  const sections = document.querySelectorAll('.stallion');
  const navLinks = document.querySelectorAll('.nav-links a');
  const linkMap  = {};
  navLinks.forEach(a => { linkMap[a.getAttribute('href').slice(1)] = a; });
  const navObs = new IntersectionObserver(entries => {
    entries.forEach(entry => {
      if (entry.isIntersecting) {
        navLinks.forEach(a => a.classList.remove('active'));
        const a = linkMap[entry.target.id];
        if (a) { a.classList.add('active'); a.scrollIntoView({ block: 'nearest', inline: 'center', behavior: 'smooth' }); }
      }
    });
  }, { threshold: 0.12, rootMargin: '-58px 0px -30% 0px' });
  sections.forEach(s => navObs.observe(s));

  // ── Back-to-top button ──
  const topBtn = document.querySelector('.back-to-top');
  window.addEventListener('scroll', () => { topBtn.classList.toggle('visible', window.scrollY > 400); }, { passive: true });

  const isMobile = () => window.innerWidth <= 600;

  // ── Collapsible sections — lazy init so hidden cards don't get max-height:0 ──
  function initCollapsibles(container) {
    container.querySelectorAll('.content-block').forEach(block => {
      if (block.dataset.cinit) return;
      block.dataset.cinit = '1';
      const title = block.querySelector('.block-title');
      const body  = block.querySelector('.block-body');
      if (!title || !body) return;
      // Start fully open (none avoids scrollHeight=0 for lazy-loaded images)
      body.style.maxHeight = 'none';
      title.addEventListener('click', () => {
        if (block.classList.contains('collapsed')) {
          block.classList.remove('collapsed');
          body.style.maxHeight = body.scrollHeight + 'px';
          body.addEventListener('transitionend', () => {
            if (!block.classList.contains('collapsed')) body.style.maxHeight = 'none';
          }, { once: true });
        } else {
          // Pin current height so CSS transition has a start value
          body.style.maxHeight = body.scrollHeight + 'px';
          requestAnimationFrame(() => requestAnimationFrame(() => {
            block.classList.add('collapsed');
            body.style.maxHeight = '0';
          }));
        }
      });
    });
  }

  // ── Mobile accordion ── (always attach; header is display:none on desktop)
  document.querySelectorAll('.accord-header').forEach(header => {
    header.addEventListener('click', () => {
      const section = header.parentElement;
      const isOpen  = section.classList.contains('accord-open');
      document.querySelectorAll('.stallion.accord-open').forEach(s => {
        s.classList.remove('accord-open');
        s.querySelector('.accord-header').setAttribute('aria-expanded', 'false');
        const sb = s.querySelector('.sticky-bar');
        if (sb) sb.classList.remove('sb-visible');
      });
      if (!isOpen) {
        section.classList.add('accord-open');
        header.setAttribute('aria-expanded', 'true');
        initCollapsibles(section);   // init now that card is visible
        initStickyBar(section);
        setTimeout(() => section.scrollIntoView({ behavior: 'smooth', block: 'start' }), 50);
      }
    });
  });

  // On desktop, init collapsibles immediately (cards already visible)
  if (!isMobile()) {
    initCollapsibles(document);
  }

  // ── Sticky name bar ──
  function initStickyBar(section) {
    const hdr = section.querySelector('.stallion-header');
    const bar = section.querySelector('.sticky-bar');
    if (!hdr || !bar || bar.dataset.sbinit) return;
    bar.dataset.sbinit = '1';
    new IntersectionObserver(([e]) => {
      const scrolledPast = !e.isIntersecting && e.boundingClientRect.top < 0;
      const open = !isMobile() || section.classList.contains('accord-open');
      bar.classList.toggle('sb-visible', scrolledPast && open);
    }, { rootMargin: '-55px 0px 0px 0px' }).observe(hdr);
  }
  if (!isMobile()) {
    document.querySelectorAll('.stallion').forEach(initStickyBar);
  }

  // ── Card entrance animation (desktop only) ──
  if (!isMobile()) {
    const cards = document.querySelectorAll('.stallion-card');
    cards.forEach(c => c.classList.add('card-anim'));
    const cardObs = new IntersectionObserver(entries => {
      entries.forEach(e => { if (e.isIntersecting) { e.target.classList.add('visible'); cardObs.unobserve(e.target); } });
    }, { threshold: 0.04 });
    cards.forEach(c => cardObs.observe(c));
  }

  // ── Share button ──
  function copyText(text) {
    if (navigator.clipboard && window.isSecureContext) return navigator.clipboard.writeText(text);
    const el = Object.assign(document.createElement('textarea'), { value: text });
    Object.assign(el.style, { position: 'fixed', opacity: '0' });
    document.body.appendChild(el); el.focus(); el.select();
    try { document.execCommand('copy'); } catch(e) {}
    document.body.removeChild(el);
    return Promise.resolve();
  }
  function showToast(msg) {
    const t = document.getElementById('toast');
    t.textContent = msg; t.classList.add('show');
    setTimeout(() => t.classList.remove('show'), 2200);
  }
  document.querySelectorAll('.share-btn').forEach(btn => {
    btn.addEventListener('click', e => {
      e.stopPropagation();
      const url = window.location.href.split('#')[0] + '#' + btn.dataset.slug;
      copyText(url).then(() => showToast('🔗 Link copied!')).catch(() => showToast('Could not copy'));
    });
  });

  // ── Swipe between stallions (mobile) ──
  let _tx = 0;
  document.addEventListener('touchstart', e => { _tx = e.touches[0].clientX; }, { passive: true });
  document.addEventListener('touchend', e => {
    if (!isMobile()) return;
    const dx = e.changedTouches[0].clientX - _tx;
    if (Math.abs(dx) < 60) return;
    const all = Array.from(document.querySelectorAll('.stallion'));
    const idx = all.findIndex(s => s.classList.contains('accord-open'));
    if (idx === -1) return;
    const nxt = dx < 0 ? idx + 1 : idx - 1;
    if (nxt < 0 || nxt >= all.length) return;
    all[idx].classList.remove('accord-open');
    all[idx].querySelector('.accord-header').setAttribute('aria-expanded', 'false');
    all[nxt].classList.add('accord-open');
    all[nxt].querySelector('.accord-header').setAttribute('aria-expanded', 'true');
    initCollapsibles(all[nxt]);
    setTimeout(() => all[nxt].scrollIntoView({ behavior: 'smooth', block: 'start' }), 50);
  }, { passive: true });

  // ── Service worker (offline support) ──
  if ('serviceWorker' in navigator) {
    navigator.serviceWorker.register('sw.js').catch(() => {});
  }

})();
"""

# ── Main builder ──────────────────────────────────────────────────────────────

def embed_favicon():
    """Return a base64 data URI for D.jpg as a 64x64 PNG favicon."""
    from PIL import Image
    path = BASE / "D.jpg"
    if not path.exists():
        return None
    img = Image.open(path).resize((64, 64), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    b64 = base64.b64encode(buf.getvalue()).decode()
    return f"data:image/png;base64,{b64}"

def embed_logo():
    """Return a base64 data URI for the navbar logo, resized to 400px wide."""
    from PIL import Image
    logo_path = BASE / "Darley-logo-WOB.jpg"
    if not logo_path.exists():
        return "Darley-logo-WOB.jpg"
    img = Image.open(logo_path)
    w, h = img.size
    new_w, new_h = 400, int(h * 400 / w)
    img = img.resize((new_w, new_h), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=85)
    b64 = base64.b64encode(buf.getvalue()).decode()
    return f"data:image/jpeg;base64,{b64}"

def write_sw():
    """Write sw.js service worker for offline caching."""
    sw = """\
const CACHE = 'darley-v1';
const URLS  = ['./index.html', './'];

self.addEventListener('install', e => {
  e.waitUntil(caches.open(CACHE).then(c => c.addAll(URLS)));
  self.skipWaiting();
});

self.addEventListener('activate', e => {
  e.waitUntil(caches.keys().then(ks =>
    Promise.all(ks.filter(k => k !== CACHE).map(k => caches.delete(k)))
  ));
  self.clients.claim();
});

self.addEventListener('fetch', e => {
  e.respondWith(
    fetch(e.request).then(r => {
      caches.open(CACHE).then(c => c.put(e.request, r.clone()));
      return r;
    }).catch(() => caches.match(e.request))
  );
});
"""
    (BASE / 'sw.js').write_text(sw, encoding='utf-8')


def write_manifest():
    """Write site.webmanifest for PWA install prompt."""
    import json
    manifest = {
        "name": "Darley America Stallion Cheat Sheets",
        "short_name": "Darley Stallions",
        "start_url": "./index.html",
        "display": "standalone",
        "background_color": "#002A85",
        "theme_color": "#002A85",
        "icons": [
            {"src": "D.jpg", "sizes": "any", "type": "image/jpeg", "purpose": "any maskable"}
        ]
    }
    (BASE / 'site.webmanifest').write_text(json.dumps(manifest, indent=2), encoding='utf-8')


def generate():
    from datetime import date

    # Stallion list from PEDIGREES, sorted alphabetically
    stallion_names = sorted(PEDIGREES.keys())

    # Scrape profiles + selling points from darleyamerica.com (one fetch per stallion)
    stallions, sp_map = scrape_all_stallions(stallion_names)

    # Merge Excel profile fields (height, YOB, earnings, etc.) into scraped profiles
    xl_stallions = load_stallions_from_xl()
    PROFILE_FIELDS = ('year_foaled', 'height', 'earnings', 'entered_stud', 'first_crop_note')
    for s in stallions:
        xl = xl_stallions.get(s.get('name', ''), {})
        for f in PROFILE_FIELDS:
            if xl.get(f):   # Excel always takes precedence over scraped value
                s[f] = xl[f]

    # Fee history + sale results from DB; highlights from Excel (falls back gracefully)
    fh_map, sr_map, hl_map = load_tables()

    n           = len(stallions)
    logo_src    = embed_logo()
    favicon_src = embed_favicon()
    _d          = date.today()
    updated     = f"{_d.strftime('%B')} {_d.day}, {_d.year}"

    # Nav links
    nav_links = '\n        '.join(
        f'<a href="#{slugify(str(s.get("name","")))}">{esc(str(s.get("name","")))}</a>'
        for s in stallions
    )

    # Stallion sections
    sections = '\n'.join(
        stallion_section(
            s,
            fh_map.get(str(s.get('name', '')), []),
            sr_map.get(str(s.get('name', '')), []),
            hl_map.get(str(s.get('name', '')), []),
            sp_map.get(str(s.get('name', '')), []),
        )
        for s in stallions
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Darley America — Stallion Cheat Sheets</title>
  <meta property="og:title" content="Darley America — Stallion Cheat Sheets">
  <meta property="og:description" content="{n} stallions · Updated {updated}">
  <meta property="og:type" content="website">
  <meta name="twitter:card" content="summary">
  <meta name="theme-color" content="#002A85">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Darley Stallions">
  <link rel="manifest" href="site.webmanifest">
  {'<link rel="icon" href="' + favicon_src + '">' if favicon_src else ''}
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Source+Sans+3:ital,wght@0,300;0,400;0,600;0,700;0,900;1,400&family=Playfair+Display:wght@700;800;900&display=swap" rel="stylesheet">
  <style>{CSS}</style>
</head>
<body id="top">

  <!-- ══ NAVBAR ══ -->
  <nav id="navbar">
    <div class="nav-logo">
      <a href="#top" style="display:flex;align-items:center;height:100%;"><img src="{logo_src}" alt="Darley America"></a>
    </div>
    <div class="nav-divider"></div>
    <div class="nav-links">
        {nav_links}
    </div>
  </nav>

  <!-- ══ HERO ══ -->
  <header class="hero">
    <h1>DARLEY AMERICA</h1>
    <p>Stallion Cheat Sheets · Updated {updated}</p>
    <div class="stallion-count">{n} Stallions</div>
  </header>

  <!-- ══ STALLION SECTIONS ══ -->
  <main>
{sections}
  </main>

  <!-- ══ FOOTER ══ -->
  <footer>
    DARLEY AMERICA &nbsp;·&nbsp; Jonabell Farm, Lexington KY &nbsp;·&nbsp; darleyamerica.com
    &nbsp;·&nbsp; Confidential — Internal Use Only
  </footer>

  <!-- ══ BACK TO TOP ══ -->
  <a href="#" class="back-to-top" aria-label="Back to top">&#8679;</a>

  <!-- ══ TOAST ══ -->
  <div class="toast" id="toast"></div>

  <script>{JS}</script>
</body>
</html>"""

    write_sw()
    write_manifest()
    OUT.write_text(html, encoding='utf-8')
    size_kb = OUT.stat().st_size // 1024
    print(f"Saved: {OUT}  ({size_kb} KB)")
    print(f"Open in your browser:  {OUT}")
    _git_push()

def _git_push():
    """Stage changed output files, commit with today's date, and push to GitHub."""
    import subprocess
    commit_msg = f"Auto-update: {date.today().strftime('%Y-%m-%d')}"
    # Files to stage (generated outputs only — not source files)
    files_to_stage = ['index.html', 'site.webmanifest', 'sw.js']
    try:
        subprocess.run(['git', 'add'] + files_to_stage, cwd=BASE, check=True)
        # Only commit if there are staged changes
        result = subprocess.run(
            ['git', 'diff', '--cached', '--quiet'],
            cwd=BASE
        )
        if result.returncode == 0:
            print("Git: no changes to commit.")
            return
        subprocess.run(['git', 'commit', '-m', commit_msg], cwd=BASE, check=True)
        subprocess.run(['git', 'push'],                      cwd=BASE, check=True)
        print(f"Git: committed and pushed — '{commit_msg}'")
    except subprocess.CalledProcessError as e:
        print(f"Git: push failed ({e}) — site saved locally but not pushed.")
    except FileNotFoundError:
        print("Git: 'git' command not found — skipping push.")

if __name__ == '__main__':
    generate()
